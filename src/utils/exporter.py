import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from src.models.shift import SHIFT_DEFINITIONS

def to_excel(df: pd.DataFrame, year: int, month: int) -> bytes:
    """
    Converts a pandas DataFrame to an Excel file in bytes, with enhanced formatting and stats.
    df lines: Employee Names. Columns: Day strings ("01", "02"...). Values: Shift Codes.
    """
    output = BytesIO()
    
    # 1. Calculate Stats
    # We need to add columns: "Debito Os." (Target), "Ore Eff.", "Saldo"
    
    # Add Stats Columns to a copy
    df_export = df.copy()
    
    from src.utils.holidays import get_italian_holidays
    from datetime import date
    holidays = get_italian_holidays(year)
    
    # Helper to calc hours
    def calc_hours(row):
        total = 0.0
        # row index usually corresponds to days ("01", "02", etc.)
        for col_name, shift_code in row.items():
            if str(col_name).isdigit(): # Ensure it's a day column
                day = int(col_name)
                curr_date = date(year, month, day)
                is_sunday = (curr_date.weekday() == 6)
                is_holiday = (curr_date in holidays)
                
                if shift_code in SHIFT_DEFINITIONS:
                    shift = SHIFT_DEFINITIONS[shift_code]
                    
                    # Se è un'assenza su giorno festivo/domenica, conta 0
                    if shift.is_absence and (is_sunday or is_holiday):
                        total += 0.0
                    else:
                        total += shift.weight
        return total

    # Calculate "Ore Effettive"
    df_export['Ore Eff.'] = df.apply(calc_hours, axis=1)
    
    # Calculate "Debito Orario" (Target) using precise Holiday Logic
    from src.utils.holidays import get_monthly_target_hours
    
    monthly_target = get_monthly_target_hours(year, month)
    
    df_export['Debito'] = monthly_target
    df_export['Saldo'] = df_export['Ore Eff.'] - df_export['Debito']
    
    # Reorder: Name (Index), Days..., Debito, Ore Eff, Saldo
    # Pandas to_excel puts Index first.
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=True, sheet_name='Turni')
        
        # Formatting
        workbook = writer.book
        worksheet = writer.sheets['Turni']
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Iterate and apply
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Header row
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Auto-width
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
    return output.getvalue()
